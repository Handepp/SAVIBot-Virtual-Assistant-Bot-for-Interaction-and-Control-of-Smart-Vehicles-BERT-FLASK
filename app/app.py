from transformers import BertTokenizer
from transformers import TFBertForSequenceClassification
import tensorflow as tf
from keras.models import load_model
from flask import Flask,render_template,request,jsonify,Response,redirect, url_for
from joblib import load
from openpyxl import Workbook, load_workbook
import random
import numpy as np
import time
import serial
import cv2
import threading
from datetime import datetime
import locale
locale.setlocale(locale.LC_TIME, "id_ID") 
import re
from init import *
from flask_basicauth import BasicAuth

import warnings
warnings.filterwarnings("ignore")

#Deklarasi variabel yang dibutuhkan diawal
camera = cv2.VideoCapture(0)
status_wahana = "SAVI ON"
k=0
k2=0
lat_tambah=0
longi_tambah=0
response_tag=''

waktu = datetime.now()
i=0
j=1
i2=0
j2=1
Respons_time=0
ws = None
ws2 = None
gerak = ""

#------------------------------------------------ FLASK ---------------------------------------------------------#
# Buat variabel flask
app   = Flask(__name__, static_url_path='/static')
app_2= Flask(__name__, static_url_path='/static')

# Membuat autentikasi
app.config['BASIC_AUTH_USERNAME'] = 'savi_unpad'
app.config['BASIC_AUTH_PASSWORD'] = 'savi123'
app.config['BASIC_AUTH_FORCE'] = True  
basic_auth = BasicAuth(app)

# Buat rute untuk template HTML dari flask
# Website publik
@app_2.route("/")
def home():
    return render_template("chatbot.html", data=data)

# Website kendali
@app.route("/")
@basic_auth.required
def home():
    return render_template("SAVI.html", data=data)

# [Routing untuk API]	
# Website kendali
@app.route("/get")
def apiDeteksi():
    # Buat variabel menjadi global
    global start_time
    global predicted_class_probability
    global response_tag
    global status_wahana 
    global data_peta
    global respons
    global k
    global lat_tambah
    global longi_tambah

    # Buat untuk respon time
    start_time = time.time()

    # Terima input dari form HTML dan JS
    chat = request.args.get('prediction_input')

    # Proses BERT
    prechat = text_preprocessing_process(chat)
    input_text_tokenized = bert_tokenizer.encode(prechat,
                                             truncation=True,
                                             padding='max_length',
                                             return_tensors='tf')
    
    bert_predict = bert_load_model(input_text_tokenized)          # Lakukan prediksi pada input
    bert_predict = tf.nn.softmax(bert_predict[0], axis=-1)         # Softmax function untuk mendapatkan hasil klasifikasi
    output = tf.argmax(bert_predict, axis=1) # Label yang terprediksi
    predicted_class_probability = bert_predict[0][output[0]].numpy() # Persentase prediksi input
    
   
    print(f'predict bert softmax:  {bert_predict}')
    print(f'output:  {output}')
    print(f'predicted class probability: {predicted_class_probability:.2%}')
    
    # Print 5 label dengan persentase tertinggi
    class_probabilities = bert_predict[0].numpy()
    label_names = le.classes_
    # Simpan pasangan label dan probabilitas dalam daftar
    predictions = [(label, probability) for label, probability in zip(label_names, class_probabilities)]

    # Urutkan daftar berdasarkan probabilitas (dalam urutan menurun)
    predictions.sort(key=lambda x: x[1], reverse=True)

    # Cetak hanya 5 nilai teratas
    top_5_predictions = predictions[:5]
    for label, probability in top_5_predictions:
        print(f'Predicted Probability for {label}: {probability:.2%}')
    predicted_label = label_names[output[0]]

    print(f'Predicted Label: {predicted_label}')

    # Data dummy gps agar latitude longitude berpindah
    lat_tambah = lat_tambah - 0.001
    longi_tambah = longi_tambah + 0.001

    response_tag = le.inverse_transform([output])[0] # Label yang terprediksi
    respons = random.choice(responses[response_tag]) # Respons yang diberikan chatbot

    # Jika prediksi input dibawah 40%
    if predicted_class_probability <= 0.4:
        response_tag = None
        respons = "Maaf, saya tidak dirancang untuk menjawab interaksi tersebut"

    # Respons untuk lokasi GPS
    if response_tag == 'Lokasi':
        try:
            arduino.write(str.encode('{"informasi":"gps"}'))
            data_gps = arduino.readline().decode().strip()
            print(data_gps)
            print(lat_tambah)
            values = data_gps.split(",")
            data_lat = float(values[0]) + lat_tambah
            data_longi = float(values[1]) + longi_tambah
            print(data_lat)
            data_peta = {
                'lat': data_lat,
                'longi': data_longi,
                }

        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan sistem."  
            data_peta = {
                'lat': None,
                'longi': None,
                }
            response_tag = None       
    
    # Respons untuk photo
    if response_tag == 'Photo':
        
        ret,frame = camera.read()

        while(True):
            cv2.imwrite(f'app\static\screenshot\img{k}.jpg',frame)
            k = k + 1
            break

    if response_tag == 'Hidup':
        status_wahana ='SAVI ON'
    if response_tag == 'Mati':
        status_wahana ='SAVI OFF'
            

    return action_if(response_tag,respons,start_time,chat), excel_wahana(chat,response_tag,predicted_class_probability)

'''
@app.route("/kendali/get")
def apiDeteksi_kendali():
    global start_time
    global response_tag
    global status_wahana 
    global data_peta
    global respons
    global k
    global lat_tambah
    global longi_tambah
    start_time = time.time()
    chat = request.args.get('prediction_input_kendali')
    prechat = text_preprocessing_process(chat)
    input_text_tokenized = bert_tokenizer.encode(prechat,
                                             truncation=True,
                                             padding='max_length',
                                             return_tensors='tf')

    bert_predict = bert_load_model(input_text_tokenized)          # Lakukan prediksi
    bert_predict = tf.nn.softmax(bert_predict[0], axis=-1)         # Softmax function untuk mendapatkan hasil klasifikasi
    output = tf.argmax(bert_predict, axis=1)
    global predicted_class_probability
    predicted_class_probability = bert_predict[0][output[0]].numpy()

    print(f'predict bert {bert_predict}')
    print(f'output {output}')
    print(f'predicted class probability: {predicted_class_probability:.2%}')
    
    class_probabilities = bert_predict[0].numpy()
    label_names = le.classes_
    # Simpan pasangan label dan probabilitas dalam daftar
    predictions = [(label, probability) for label, probability in zip(label_names, class_probabilities)]

    # Urutkan daftar berdasarkan probabilitas (dalam urutan menurun)
    predictions.sort(key=lambda x: x[1], reverse=True)

    # Cetak hanya 5 nilai teratas
    top_5_predictions = predictions[:5]
    for label, probability in top_5_predictions:
        print(f'Predicted Probability for {label}: {probability:.2%}')
    predicted_label = label_names[output[0]]

    print(f'Predicted Label: {predicted_label}')



    lat_tambah = lat_tambah - 0.001
    longi_tambah = longi_tambah + 0.001
    
    response_tag = le.inverse_transform([output])[0]
    respons = random.choice(responses[response_tag])

    if predicted_class_probability <= 0.4:
        response_tag = None
        respons = "Maaf, saya tidak dirancang untuk menjawab interaksi tersebut"   

    if response_tag == 'Hidup':
        status_wahana ='SAVI ON'
    if response_tag == 'Mati':
        status_wahana ='SAVI OFF'
            

    return action_if(response_tag,respons,start_time,chat), excel_wahana(chat,response_tag,predicted_class_probability)'''

# Website publik
@app_2.route("/get")
def apiDeteksi():
    global start_time
    global response_tag
    global respons
    global data_peta
    global k2
    global lat_tambah
    global longi_tambah

    start_time = time.time()
    chat = request.args.get('prediction_input')
    prechat = text_preprocessing_process(chat)
    input_text_tokenized = bert_tokenizer.encode(prechat,
                                             truncation=True,
                                             padding='max_length',
                                             return_tensors='tf')

    bert_predict = bert_load_model(input_text_tokenized)          # Lakukan prediksi
    bert_predict = tf.nn.softmax(bert_predict[0], axis=-1)         # Softmax function untuk mendapatkan hasil klasifikasi
    output = tf.argmax(bert_predict, axis=1)
    global predicted_class_probability
    predicted_class_probability = bert_predict[0][output[0]].numpy()

    print(f'predict bert {bert_predict}')
    print(f'output {output}')
    print(f'predicted class probability: {predicted_class_probability:.2%}')
    
    class_probabilities = bert_predict[0].numpy()
    label_names = le.classes_
    # Simpan pasangan label dan probabilitas dalam daftar
    predictions = [(label, probability) for label, probability in zip(label_names, class_probabilities)]

    # Urutkan daftar berdasarkan probabilitas (dalam urutan menurun)
    predictions.sort(key=lambda x: x[1], reverse=True)

    # Cetak hanya 5 nilai teratas
    top_5_predictions = predictions[:5]
    for label, probability in top_5_predictions:
        print(f'Predicted Probability for {label}: {probability:.2%}')
    predicted_label = label_names[output[0]]

    print(f'Predicted Label: {predicted_label}')

    lat_tambah = lat_tambah - 0.001
    longi_tambah = longi_tambah + 0.001
    
    response_tag = le.inverse_transform([output])[0]
    respons = random.choice(responses[response_tag])

    if predicted_class_probability <= 0.4:
        response_tag = None
        respons = "Maaf, saya tidak dirancang untuk menjawab interaksi tersebut"

    if response_tag == 'Lokasi':
        try:
            arduino.write(str.encode('{"informasi":"gps"}'))
            data_gps = arduino.readline().decode().strip()
            print(data_gps)
            print(lat_tambah)
            values = data_gps.split(",")
            data_lat = float(values[0]) + lat_tambah
            data_longi = float(values[1]) + longi_tambah
            print(data_lat)
            data_peta = {
                'lat': data_lat,
                'longi': data_longi,
                }

        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan sistem."  
            data_peta = {
                'lat': None,
                'longi': None,
                }
            response_tag = None     
    
    if response_tag == 'Photo':
        
        ret,frame = camera.read()

        while(True):
            cv2.imwrite(f'app\static\screenshot_2\img{k2}.jpg',frame)
            k2 = k2 + 1
            break
    return action_if_publik(response_tag,respons,start_time), excel_wahana_publik(chat,response_tag,predicted_class_probability)

# Buat workbook excel
try:
    wb = load_workbook("Data_Excel\SAVIBot_Kendali_wahana.xlsx")
    wb2 = load_workbook("Data_Excel\SAVIBot_Fitur.xlsx")
except FileNotFoundError:
    wb = Workbook()
    wb2 = Workbook()

# Fungsi bagian hari
def part_hari(h):
    if 5 <= h <= 11:
        return "Pagi"
    elif 12 <= h <= 16 :
        return "Siang"
    elif 16 <= h <= 18 :
        return "Sore"
    else:
        return "Malam"

part = part_hari(datetime.now().hour)

# Mendapatkan respons time
def time_chatbot(start_time):
    global Respons_time
    end_time = time.time()
    Respons_time = end_time - start_time
    print("Waktu Respons:", Respons_time, 'detik')
    return Respons_time

# Membulatkan nilai
def Pembulatan(value):
    value1 = float(value)
    value1 = round(value1,0)
    value1 = int(value1)
    value1 = str(value1)
    return value1

#Replace (,) menjadi (.) dan seleksi angka pada kalimat
def ex_number(text):
  text = text.replace(",", ".") 
  temp = re.findall(r'\d+\.\d+|\d+', text)
  res = list(map(float, temp))

  return res

# Aksi yang diberikan chatbot berdasarkan hasil prediksi input (Kendali)
def action_if(label,res,waktu,chat):
    #global relay_data
    #global SAVIstatus
    global respons
    global gerak

    if(label == 'Hidup' ):
        try:
            arduino.write(str.encode('{"control":"hidup"}'))
            res_arduino = arduino.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = ""
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
    
    if(label == 'Mati' ):
        try:
            arduino_2.write(str.encode('{"control":"reset"}'))
            arduino.write(str.encode('{"control":"mati"}'))
            res_arduino = arduino.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = ""
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons       
    
    if(label == 'Maju' ):
        try:
            #arduino_2.write(str.encode('{"control":"maju"}'))
            arduino.write(str.encode('{"control":"maju"}'))
            res_arduino = arduino.readline().decode().strip()
            #res_arduino_2 = arduino_2.readline().decode().strip()
            time_chatbot(waktu)
            #speak(respons)
            respons = res
            gerak = "Maju"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
    
    if (label == 'Mundur'):
        try:
            arduino_2.write(str.encode('{"control":"mundur"}'))
            arduino.write(str.encode('{"control":"mundur"}'))
            res_arduino = arduino.readline().decode().strip()
            #res_arduino_2 = arduino_2.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = "Mundur"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
    
    if (label == 'Stop'):
        try:
            arduino_2.write(str.encode('{"control":"stop"}'))
            arduino.write(str.encode('{"control":"stop"}'))
            res_arduino = arduino.readline().decode().strip()
            #res_arduino_2 = arduino_2.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = "Berhenti"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
        
    if (label == 'Kanan'):
        try:
            arduino.write(str.encode('{"control":"stop"}'))
            time.sleep(1.5)
            arduino_2.write(str.encode('{"control":"kanan"}'))
            time.sleep(1.5)
            arduino.write(str.encode('{"control":"kanan"}'))
            res_arduino = arduino.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = "Belok kanan"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
    
    if (label == 'Kiri'):
        try:
            arduino.write(str.encode('{"control":"stop"}'))
            time.sleep(1.5)
            arduino_2.write(str.encode('{"control":"kiri"}'))
            time.sleep(1.5)
            arduino.write(str.encode('{"control":"kiri"}'))
            res_arduino = arduino.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = "Belok kiri"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
        
    if (label == 'Putar'):
        try:
            arduino.write(str.encode('{"control":"stop"}'))
            time.sleep(1.5)
            arduino_2.write(str.encode('{"control":"putar_balik"}'))
            time.sleep(1.5)
            arduino.write(str.encode('{"control":"putar_balik"}'))
            res_arduino = arduino.readline().decode().strip()
            time_chatbot(waktu)
            respons = res
            gerak = "Putar balik"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons

    if (label == 'Speed'):
        #Input kecepatan dalam m/s
        kec=ex_number(chat)
        kec_send = str(kec[0])
        print(kec_send)
        try:
            print('cek1')
            arduino.write(str.encode('{"control":"speed","velo1":"' + kec_send + '","velo2":"' + kec_send + '"}'))
            #time.sleep(2)  # Beri waktu untuk membuka koneksi
            #arduino.write(kec_send.encode())
            res_arduino = arduino.readline().decode().strip()
            print(res_arduino)
            print(kec_send)
            values = res_arduino.split(",")
            kec_now = float(values[0])
            kec_now = str(kec_now)
            kec_now = kec_now.replace(".", ",")
            respons = res+ " " + kec_now + ' ' + "m/s"
            print(respons)
            gerak = f"Kecepatan menjadi {kec_now}m/s"
            time_chatbot(waktu)
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
        
    if(label == 'Infospeed'):
        try:
            arduino.write(str.encode('{"informasi":"infospeed"}'))
            res_arduino = arduino.readline().decode('utf-8').strip()
            values = res_arduino.split(",")
            kec_now = float(values[0])
            print(kec_now)
            kec_now=str(kec_now)
            kec_now = kec_now.replace(".", ",")
            respons = res+ " " + "0,5"+ ' ' + "m/s"
            gerak = ""
            time_chatbot(waktu)
            return  respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons



    if(label == 'Suhu'):
        try:
            arduino.write(str.encode('{"informasi":"temp"}'))
            res_arduino = arduino.readline().decode().strip() # 27.90,1001
            print(res_arduino)
            values = res_arduino.split(",") #[27.90, 1001]
            print(values)
            data_suhu = float(values[0])
            print(data_suhu)
            data_suhu=str(data_suhu)
            data_suhu = data_suhu.replace(".", ",")
            respons = res + " " + data_suhu + " " + "derajat celcius"
            time_chatbot(waktu)
            gerak = ""
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons

    if(label == 'Hump'):
        try:
            arduino.write(str.encode('{"informasi":"hum"}'))
            res_arduino = arduino.readline().decode().strip()
            values = res_arduino.split(",")
            data_hump = float(values[0])
            print(data_hump)
            data_hump = Pembulatan(data_hump)
            time_chatbot(waktu)
            respons = res + " " + data_hump + " " + "RH"
            gerak = ""
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons

    if(label == 'Jam'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%H %M") + ' ' + part
        gerak = ""
        return respons

    if(label == 'Hari'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%A")
        gerak = ""
        return respons

    if(label == 'Tanggal'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%d %B %Y")
        gerak = ""
        return respons
    
    if(label == 'Usia'):
        time_chatbot(waktu)
        respons = res + '.' + ' ' + umur()
        gerak = ""
        return respons
    
    if(label == 'Lokasi'):
        time_chatbot(waktu)
        gerak = ""
    
    if(label == 'Baterai'):
        try:
            arduino.write(str.encode('{"informasi":"baterai"}'))
            res_arduino = arduino.readline().decode().strip()
            values = res_arduino.split(",")
            data_baterai = float(values[0])
            print(data_baterai)
            data_baterai=str(data_baterai)
            data_baterai = data_baterai.replace(".", ",")
            respons = res + ' ' + data_baterai +'%'
            time_chatbot(waktu) 
            gerak = ""
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            gerak = ""
            return respons
    
    else:
        time_chatbot(waktu)
        respons = res
        gerak = ""
        return respons

# Masukan hasil ke file excel (Publik)
def excel_wahana(input, y_pred, pred_prob):
    global i
    global j
    global Respons_time
    
    Nama = 'test buat data3'
    ws = None
    for sheet in wb.sheetnames:
        if sheet == Nama:
            ws = wb[sheet]
            break

    # Jika sheet belum ada, buatlah sheet baru
    if ws is None:
        ws = wb.create_sheet(Nama)

    ws['A1'] = 'Uji ke'
    ws['B1'] = 'Input Kalimat'
    ws['C1'] = 'Respon SAVIBot'
    ws['D1'] = 'Respons Time'
    ws['E1'] = 'Pergerakan'
    ws['F1'] = 'Persentase Prediksi'
    ws['G1'] = 'Label prediksi'

    ws.cell(row=2+i, column=1).value = j
    ws.cell(row=2+i, column=2).value = input
    ws.cell(row=2+i, column=3).value = respons
    ws.cell(row=2+i, column=4).value = Respons_time
    ws.cell(row=2+i, column=5).value = gerak
    ws.cell(row=2+i, column=6).value = pred_prob*100
    ws.cell(row=2+i, column=7).value = y_pred

    i+=1
    j+=1
    # Save the file
    wb.save("Data_Excel\SAVIBot_Kendali_wahana.xlsx")

# Mendapatkan informasi waktu
def get_time(time):
    waktu = datetime.now()
    waktu = waktu.strftime(f"{time}")
    return waktu

# Mendapatkan informasi umur
def umur():
    tanggal_lahir = datetime(2023, 7, 16)
    selisih_tahun = waktu.year - tanggal_lahir.year
    selisih_bulan = waktu.month - tanggal_lahir.month
    if tanggal_lahir.month > waktu.month or (tanggal_lahir.month == waktu.month and tanggal_lahir.day > waktu.day):
        selisih_tahun -= 1

    if selisih_tahun>0:
        return(f'Umur saya adalah {selisih_tahun} tahun' f' {selisih_bulan} bulan')
    else:
        return(f'Umur saya adalah {selisih_bulan} bulan')

# Aksi yang diberikan chatbot berdasarkan hasil prediksi input (Publik)
def action_if_publik(label,res,waktu):
    #global relay_data
    #global SAVIstatus
    global respons
    if(label == 'Hidup' or label == 'Mati' or label == 'Maju' or label == 'Mundur' or label == 'Stop' 
       or label == 'Speed' or label == 'Kanan' or label == 'Kiri' or label == "Putar"):
        respons = "Maaf, anda tidak bisa menggunakan fitur kendali melalui chatroom ini. "
        print(respons)
        time_chatbot(waktu)
        return respons

    if(label == 'Suhu'):
        try:
            arduino.write(str.encode('{"informasi":"temp"}'))
            res_arduino = arduino.readline().decode().strip() # 27.90,1001
            print(res_arduino)
            values = res_arduino.split(",") #[27.90, 1001]
            print(values)
            data_suhu = float(values[0])
            print(data_suhu)
            data_suhu=str(data_suhu)
            data_suhu = data_suhu.replace(".", ",")
            respons = res + " " + data_suhu + " " + "derajat celcius"
            time_chatbot(waktu)
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            return respons

    if(label == 'Hump'):
        try:
            arduino.write(str.encode('{"informasi":"hum"}'))
            res_arduino = arduino.readline().decode().strip()
            values = res_arduino.split(",")
            data_hump = float(values[0])
            print(data_hump)
            data_hump = Pembulatan(data_hump)
            time_chatbot(waktu)
            respons = res + " " + data_hump + " " + "RH"
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            return respons

    if(label == 'Jam'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%H %M") + ' ' + part
        return respons

    if(label == 'Hari'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%A")
        return respons

    if(label == 'Tanggal'):
        time_chatbot(waktu)
        respons = res + ' ' + get_time("%d %B %Y")
        return respons
    
    if(label == 'Usia'):
        time_chatbot(waktu)
        respons = res + '.' + ' ' + umur()
        return respons
    
    if(label == 'Lokasi'):
        time_chatbot(waktu)
    
    if(label == 'Baterai'):
        try:
            arduino.write(str.encode('{"informasi":"baterai"}'))
            res_arduino = arduino.readline().decode().strip()
            values = res_arduino.split(",")
            data_baterai = float(values[0])
            print(data_baterai)
            data_baterai=str(data_baterai)
            data_baterai = data_baterai.replace(".", ",")
            respons = res + ' ' + data_baterai +'%'
            time_chatbot(waktu) 
            return respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            return respons
            
    
    if(label == 'Infospeed'):
        try:
            arduino.write(str.encode('{"informasi":"infospeed"}'))
            res_arduino = arduino.readline().decode('utf-8').strip()
            values = res_arduino.split(",")
            kec_now = float(values[0])
            print(kec_now)
            kec_now=str(kec_now)
            kec_now = kec_now.replace(".", ",")
            respons = res+ " " + kec_now+ ' ' + "m/s"
            time_chatbot(waktu)
            return  respons
        except:
            respons = "Maaf, saat ini SAVI sedang tidak terhubung dengan mikrokontroler."
            print(respons)
            time_chatbot(waktu)
            return respons           

    else:
        time_chatbot(waktu)
        respons = res
        return respons

# Masukan hasil ke file excel
def excel_wahana_publik(input, y_pred, pred_prob):
    global i2
    global j2
    global Respons_time
    
    Nama = "test_video"
    ws2 = None
    for sheet in wb.sheetnames:
        if sheet == Nama:
            ws2 = wb2[sheet]
            break

    # Jika sheet belum ada, buatlah sheet baru
    if ws2 is None:
        ws2 = wb2.create_sheet(Nama)

    # Cek apakah sheet dengan nama "Pengujian fitur lokasi" sudah ada

    ws2['A1'] = 'Uji ke'
    ws2['B1'] = 'Input Kalimat'
    ws2['C1'] = 'Label Terprediksi'
    ws2['D1'] = 'Respon SAVIBot'
    ws2['E1'] = 'Persentase Predik'
    ws2['F1'] = 'Respons Time'

    ws2.cell(row=2+i2, column=1).value = j2
    ws2.cell(row=2+i2, column=2).value = input
    ws2.cell(row=2+i2, column=3).value = y_pred
    ws2.cell(row=2+i2, column=4).value = respons
    ws2.cell(row=2+i2, column=5).value = pred_prob*100
    ws2.cell(row=2+i2, column=6).value = Respons_time

    i2+=1
    j2+=1
    # Save the file
    wb2.save("Data_Excel\SAVIBot_Publik.xlsx")

# Route ke js
@app.route("/image")
def image():
    return jsonify(k=k)

@app.route("/tag")
def tag():
    return jsonify(response_tag=response_tag)

@app.route("/kendali/tag")
def tagkendali():
    return jsonify(response_tag=response_tag)

@app.route("/peta")
def peta():
    return jsonify(data_peta)

@app_2.route("/image")
def image():
    return jsonify(k2=k2)

@app_2.route("/tag")
def tag():
    return jsonify(response_tag=response_tag)

@app_2.route("/peta")
def peta():
    return jsonify(data_peta)

@app.route('/status')
def get_status():
    #relay_data = arduino.readline().decode('utf-8').strip()
    #arduino.close()
    #print(relay_data)
    return jsonify({'status': status_wahana })

# Run app didua port lokal host
def run_app_1():
    app.run(host="localhost", port=5000, debug=False)

def run_app_2():
    app_2.run(host="localhost", port=8080, debug=False)

if __name__ == '__main__':
    # Koneksikan serial python dengan arduino
    try:
        arduino = serial.Serial('COM3',115200)
        print('Arduino sudah tersambung dengan SAVI')
        for i in range(2):
            res_arduino = arduino.readline().decode().strip()
            print(res_arduino)
    except:
        print('Arduino tidak tersambung')

    try:
        arduino_2 = serial.Serial('COM15',115200)
        print('Arduino 2 sudah tersambung dengan SAVI')
        
    except:
        print("Arduino 2 tidak tersambung dengan SAVI")
   

    #Pretrained Model
    PRE_TRAINED_MODEL = 'indobenchmark/indobert-base-p2'

    #Load tokenizer dari pretrained model
    bert_tokenizer = BertTokenizer.from_pretrained(PRE_TRAINED_MODEL)

    # Load hasil fine-tuning
    bert_load_model = TFBertForSequenceClassification.from_pretrained(PRE_TRAINED_MODEL, num_labels=43)
    bert_load_model.load_weights('Model\Bert_16_5e_30.h5')

    #Deploy di localhost
    thread_app_1 = threading.Thread(target=run_app_1)
    thread_app_1.start()

    thread_app_2 = threading.Thread(target=run_app_2)
    thread_app_2.start()





    

    

